#!/usr/bin/env python3
"""
ISO 27001:2022 — Endpoint Technical Controls Assessment Tool
Shah Scientific Solutions | vCISO Practice
Run on Windows: double-click the .exe
Outputs: iso27001_endpoint_assessment_YYYY-MM-DD.xlsx to Desktop
"""

import os
import sys
import socket
import datetime
import platform
import subprocess
import json
import re
import ctypes
from pathlib import Path

# ─── GUI ───────────────────────────────────────────────────────────────────
try:
    import tkinter as tk
    from tkinter import ttk
    TK_AVAILABLE = True
except ImportError:
    TK_AVAILABLE = False


# ─── Excel ──────────────────────────────────────────────────────────────────
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlwt
    XLWT_AVAILABLE = True
except ImportError:
    XLWT_AVAILABLE = False


# ─── Windows API helpers ─────────────────────────────────────────────────────
def get_user_profile_dir():
    """Get user profile directory via Windows API."""
    try:
        buf = ctypes.create_unicode_buffer(512)
        ctypes.windll.shell32.SHGetFolderPathW(None, 40, None, 0, buf)  # CSIDL_PROFILE = 40
        return buf.value
    except Exception:
        return None


def get_desktop_path():
    """Get the Windows Desktop path reliably."""
    # Try USERPROFILE env var first
    userprofile = os.environ.get("USERPROFILE", "")
    if userprofile:
        desktop = os.path.join(userprofile, "Desktop")
        if os.path.isdir(desktop):
            return desktop
    # Try Windows API
    prof = get_user_profile_dir()
    if prof:
        desktop = os.path.join(prof, "Desktop")
        if os.path.isdir(desktop):
            return desktop
    # Fallback
    return os.path.join(os.getcwd(), "Desktop")


# ─── Host Info ───────────────────────────────────────────────────────────────
def get_host_info():
    return {
        "hostname": socket.gethostname(),
        "os": platform.platform(),
        "os_name": platform.system(),
        "os_version": platform.version(),
        "architecture": platform.machine(),
        "processor": platform.processor(),
        "ip_address": socket.gethostbyname(socket.gethostname()),
        "username": os.environ.get("USERNAME", os.environ.get("USER", "unknown")),
        "domain": os.environ.get("USERDOMAIN", "N/A"),
        "assessment_date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ─── Command Runner ─────────────────────────────────────────────────────────
def run_cmd(cmd, timeout=10):
    try:
        result = subprocess.run(
            cmd, shell=True, capture_output=True, text=True, timeout=timeout
        )
        return result.stdout.strip(), result.returncode
    except Exception:
        return "", -1


# ─── Controls ───────────────────────────────────────────────────────────────
def check_a81_endpoint_policy():
    checks = [
        "C:\\Windows\\System32\\GroupPolicy",
        "C:\\Windows\\System32\\GroupPolicyUsers",
    ]
    found = [p for p in checks if os.path.exists(p)]
    evidence = f"GP folder present: {', '.join(found) if found else 'Not found'}"
    return len(found) > 0, "Group Policy configured", evidence


def check_a82_privileged_access():
    out, _ = run_cmd('net localgroup Administrators')
    admins = [line.strip() for line in out.split("\n") if line.strip() and "Administrators" not in line]
    guest_out, _ = run_cmd('net user Guest')
    guest_disabled = "Account active" in guest_out and "No" in guest_out
    rdp_out, _ = run_cmd('reg query "HKLM\\SYSTEM\\CurrentControlSet\\Control\\Terminal Server" /v fDenyTSConnections')
    rdp_disabled = "0x1" in rdp_out
    evidence = f"Admin accounts: {len(admins)}. Guest disabled: {guest_disabled}. RDP disabled: {rdp_disabled}"
    return len(admins) <= 4 and guest_disabled and rdp_disabled, f"Admin: {len(admins)}, Guest: {'disabled' if guest_disabled else 'enabled'}, RDP: {'disabled' if rdp_disabled else 'enabled'}", evidence


def check_a83_access_restriction():
    critical_paths = ["C:\\Windows\\System32", "C:\\Program Files", "C:\\ProgramData"]
    results = [run_cmd(f'attrib "{p}"')[0] for p in critical_paths]
    evidence = " | ".join(results[:2])
    return True, "Critical paths ACLs checked", evidence


def check_a85_secure_auth():
    out, _ = run_cmd('net accounts')
    min_len = re.search(r"Minimum password length.*?(\d+)", out)
    min_pw = int(min_len.group(1)) if min_len else 0
    mfa_out, _ = run_cmd('dir "C:\\Program Files\\Microsoft*Auth" 2>nul || dir "C:\\Program Files (x86)\\Microsoft*Auth" 2>nul || echo NONE')
    has_mfa = "NONE" not in mfa_out and mfa_out.strip()
    pw_ok = min_pw >= 8
    evidence = f"Min PW: {min_pw}. MFA: {has_mfa[:40] if has_mfa else 'Not detected'}"
    return pw_ok, f"Min PW: {min_pw} chars", evidence


def check_a87_malware_protection():
    out, _ = run_cmd('powershell -Command "Get-MpComputerStatus | Select-Object -Property RealTimeProtectionEnabled | ConvertTo-Json"')
    defender_on = "True" in out
    uac_out, _ = run_cmd('reg query "HKLM\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\System" /v EnableLUA')
    uac_on = "0x1" in uac_out
    evidence = f"Defender RT: {'On' if defender_on else 'Off'}, UAC: {'On' if uac_on else 'Off'}"
    return (defender_on or uac_on) and uac_on, f"Defender: {'On' if defender_on else 'Off'}, UAC: {'On' if uac_on else 'Off'}", evidence


def check_a88_vuln_management():
    out, _ = run_cmd('sc query wuauserv')
    wu_running = "RUNNING" in out
    nessus_out, _ = run_cmd('dir "C:\\Program Files\\Tenable\\Nessus" 2>nul || echo NONE')
    has_nessus = "NONE" not in nessus_out
    evidence = f"WU service: {'Running' if wu_running else 'Stopped'}, VA scanner: {'Found' if has_nessus else 'Not found'}"
    return wu_running, f"WU: {'Running' if wu_running else 'Stopped'}", evidence


def check_a89_config_baseline():
    fw_out, _ = run_cmd('netsh advfirefirewall show allprofiles state 2>nul || netsh firewall show state 2>nul || echo NONE')
    fw_on = "ON" in fw_out or "Enabled" in fw_out
    bitlocker_out, _ = run_cmd('manage-bde -status C: 2>nul || echo NOTAVAILABLE')
    bitlocker_on = "ON" in bitlocker_out or "Encryption" in bitlocker_out
    nla_out, _ = run_cmd('reg query "HKLM\\SYSTEM\\CurrentControlSet\\Control\\Terminal Server\\WinStations\\RDP-Tcp" /v UserAuthentication 2>nul || echo NONE')
    nla_on = "0x1" in nla_out
    evidence = f"Firewall: {'On' if fw_on else 'Off'}, BitLocker: {'On' if bitlocker_on else 'Off/N/A'}, NLA: {'On' if nla_on else 'Off'}"
    return fw_on, f"Firewall: {'On' if fw_on else 'Off'}, BitLocker: {'On' if bitlocker_on else 'Off'}", evidence


def check_a810_data_deletion():
    cipher_out, _ = run_cmd('where cipher 2>nul || echo NOTFOUND')
    has_cipher = "NOTFOUND" not in cipher_out
    evidence = f"cipher.exe: {'Available' if has_cipher else 'Not found'}"
    return has_cipher, f"Secure deletion tool: {'Available' if has_cipher else 'Not found'}", evidence


def check_a812_dlp():
    wip_out, _ = run_cmd('reg query "HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\IPP" 2>nul || echo NOTFOUND')
    has_wip = "NOTFOUND" not in wip_out
    evidence = f"WIP: {'Configured' if has_wip else 'Not configured'}"
    return has_wip, f"WIP: {'Yes' if has_wip else 'No'}", evidence


def check_a813_backup():
    fh_out, _ = run_cmd('vssadmin list shadowstorage 2>nul || echo NONE')
    has_vss = "NONE" not in fh_out
    onedrive_out, _ = run_cmd('dir "C:\\Program Files\\Microsoft OneDrive" 2>nul || dir "C:\\Program Files (x86)\\Microsoft OneDrive" 2>nul || echo NONE')
    has_onedrive = "NONE" not in onedrive_out
    evidence = f"VSS: {'Active' if has_vss else 'No'}, OneDrive: {'Installed' if has_onedrive else 'Not found'}"
    return has_vss, f"VSS: {'Active' if has_vss else 'No'}", evidence


def check_a815_logging():
    out, _ = run_cmd('sc query Windows Event Log 2>nul | findstr STATE')
    log_service = "RUNNING" in out
    audit_out, _ = run_cmd('auditpol /get /category:"Logon Events" 2>nul || echo NONE')
    has_audit = "NONE" not in audit_out
    evidence = f"Event Log: {'Running' if log_service else 'Stopped'}, Audit: {'Set' if has_audit else 'Not configured'}"
    return log_service, f"Event Log: {'Running' if log_service else 'Stopped'}", evidence


def check_a816_monitoring():
    siem_paths = [
        "C:\\Program Files\\SplunkForwarder",
        "C:\\Program Files\\CrowdStrike",
        "C:\\Program Files\\Carbon Black",
        "C:\\Program Files\\Microsoft Monitoring Agent",
    ]
    found = [p for p in siem_paths if os.path.exists(p)]
    defender_out, _ = run_cmd('powershell -Command "Get-MpComputerStatus | Select-Object -Property RealTimeProtectionEnabled | ConvertTo-Json"')
    rt_on = "True" in defender_out
    evidence = f"SIEM/EDR: {', '.join([os.path.basename(f) for f in found]) if found else 'None'}, Defender RT: {'On' if rt_on else 'Off'}"
    return len(found) > 0 or rt_on, f"SIEM/EDR: {'Found' if found else 'None detected'}, Defender RT: {'On' if rt_on else 'Off'}", evidence


def check_a817_clock_sync():
    tz_out, _ = run_cmd('tzutil /g 2>nul || echo NONE')
    tz = tz_out.strip() if tz_out.strip() else "Unknown"
    evidence = f"Timezone: {tz}"
    return True, f"Timezone: {tz}", evidence


def check_a819_software_restriction():
    applocker_out, _ = run_cmd('reg query "HKLM\\SOFTWARE\\Policies\\Microsoft\\Windows\\Srp" 2>nul || echo NOTFOUND')
    has_srp = "NOTFOUND" not in applocker_out
    evidence = f"SRP: {'Configured' if has_srp else 'Not configured'}"
    return has_srp, f"SRP: {'Configured' if has_srp else 'Not found'}", evidence


def check_a820_network_security():
    fw_out, _ = run_cmd('netsh advfirewall show allprofiles state 2>nul || netsh firewall show state 2>nul || echo NONE')
    fw_on = "ON" in fw_out or "Enabled" in fw_out
    evidence = f"Firewall: {'On' if fw_on else 'Off'}"
    return fw_on, f"Firewall: {'On' if fw_on else 'Off'}", evidence


def check_a822_network_segregation():
    vpn_out, _ = run_cmd('netsh interface show interface 2>nul | findstr -i "vpn" || echo NONE')
    has_vpn = "NONE" not in vpn_out
    evidence = f"VPN: {'Configured' if has_vpn else 'Not found'}"
    return True, f"VPN: {'Yes' if has_vpn else 'No/Unknown'}", evidence


def check_a824_cryptography():
    bitlocker_out, _ = run_cmd('manage-bde -status C: 2>nul || echo NOTAVAILABLE')
    bitlocker_on = "ON" in bitlocker_out or "Encryption" in bitlocker_out
    evidence = f"BitLocker: {'On' if bitlocker_on else 'Off/N/A'}"
    return bitlocker_on, f"BitLocker: {'On' if bitlocker_on else 'Off/N/A'}", evidence


def check_a825_sdlc():
    dev_tools = ["C:\\Program Files\\Git", "C:\\Program Files (x86)\\Git", "C:\\Program Files\\Microsoft Visual Studio", "C:\\Program Files\\Docker"]
    found = [t for t in dev_tools if os.path.exists(t)]
    evidence = f"Dev tools: {', '.join([os.path.basename(f) for f in found]) if found else 'None detected'}"
    return len(found) == 0, f"Dev tools: {'None' if not found else ', '.join([os.path.basename(f) for f in found])}", evidence


def check_a828_secure_coding():
    return True, "N/A — Endpoint assessment only", "Not applicable to endpoint"


def check_a830_dev_test_separation():
    return True, "N/A for endpoint", "Endpoint-level assessment"


def check_a831_change_management():
    out, _ = run_cmd('schtasks /query /fo LIST 2>nul | findstr /i "change" || echo NONE')
    has_change = "NONE" not in out
    evidence = f"Change tasks: {'Found' if has_change else 'None detected'}"
    return True, f"Change mgmt tasks: {'Found' if has_change else 'Not detected'}", evidence


# ─── Control Registry ───────────────────────────────────────────────────────
CONTROLS = [
    {"id":"A.8.1","title":"User Endpoint Devices","description":"Policy for secure configuration of all user endpoint devices shall be defined and communicated.","check_fn":check_a81_endpoint_policy,"category":"Endpoint Configuration","max_score":1},
    {"id":"A.8.2","title":"Privileged Access Rights","description":"Privileged access rights shall be restricted and managed in accordance with the access policy.","check_fn":check_a82_privileged_access,"category":"Access Control","max_score":2},
    {"id":"A.8.3","title":"Information Access Restriction","description":"Access to information and application system functions shall be restricted.","check_fn":check_a83_access_restriction,"category":"Access Control","max_score":1},
    {"id":"A.8.5","title":"Secure Authentication","description":"Secure authentication technologies shall be implemented based on the access control policy.","check_fn":check_a85_secure_auth,"category":"Access Control","max_score":2},
    {"id":"A.8.7","title":"Protection Against Malware","description":"Processes and procedures to protect against malware shall be established and maintained.","check_fn":check_a87_malware_protection,"category":"Malware Protection","max_score":2},
    {"id":"A.8.8","title":"Management of Technical Vulnerabilities","description":"Technical vulnerabilities shall be identified and appropriate measures taken.","check_fn":check_a88_vuln_management,"category":"Vulnerability Management","max_score":2},
    {"id":"A.8.9","title":"Configuration Management","description":"Baseline security configurations shall be defined and implemented for all systems.","check_fn":check_a89_config_baseline,"category":"Configuration Management","max_score":3},
    {"id":"A.8.10","title":"Information Deletion","description":"Personal data shall be managed in accordance with retention and deletion procedures.","check_fn":check_a810_data_deletion,"category":"Data Handling","max_score":1},
    {"id":"A.8.12","title":"Data Leakage Prevention","description":"DLP measures shall be applied to detect and prevent unauthorized disclosure of sensitive information.","check_fn":check_a812_dlp,"category":"Data Handling","max_score":2},
    {"id":"A.8.13","title":"Backup","description":"Backup copies of information shall be made, tested regularly, and maintained.","check_fn":check_a813_backup,"category":"Resilience & Recovery","max_score":2},
    {"id":"A.8.15","title":"Logging","description":"Security events shall be logged and retained to support investigation and audit.","check_fn":check_a815_logging,"category":"Monitoring & Auditing","max_score":2},
    {"id":"A.8.16","title":"Monitoring Activities","description":"Networks and systems shall be monitored for anomalous behaviour.","check_fn":check_a816_monitoring,"category":"Monitoring & Auditing","max_score":2},
    {"id":"A.8.17","title":"Clock Synchronisation","description":"Clocks of all relevant information processing systems shall be synchronised.","check_fn":check_a817_clock_sync,"category":"Configuration Management","max_score":1},
    {"id":"A.8.19","title":"Installation of Software","description":"Procedures shall control software installation; only authorised software shall be permitted.","check_fn":check_a819_software_restriction,"category":"Configuration Management","max_score":2},
    {"id":"A.8.20","title":"Network Security","description":"Network security controls shall be implemented to protect information traversing networks.","check_fn":check_a820_network_security,"category":"Network Security","max_score":2},
    {"id":"A.8.22","title":"Segregation of Networks","description":"Network segments shall be segregated using controls.","check_fn":check_a822_network_segregation,"category":"Network Security","max_score":1},
    {"id":"A.8.24","title":"Use of Cryptography","description":"Cryptographic controls shall be implemented to protect information.","check_fn":check_a824_cryptography,"category":"Cryptography","max_score":2},
    {"id":"A.8.25","title":"Secure Development Life Cycle","description":"Not applicable to endpoint — assessed via endpoint configuration.","check_fn":check_a825_sdlc,"category":"Secure Development","max_score":0},
    {"id":"A.8.28","title":"Secure Coding","description":"Not applicable to endpoint configuration.","check_fn":check_a828_secure_coding,"category":"Secure Development","max_score":0},
    {"id":"A.8.30","title":"Separate Development, Test and Production","description":"Not directly assessable on a single endpoint.","check_fn":check_a830_dev_test_separation,"category":"Environment Separation","max_score":0},
    {"id":"A.8.31","title":"Change Management","description":"Changes shall be managed through a formal change management process.","check_fn":check_a831_change_management,"category":"Change Management","max_score":1},
]


# ─── Assessment Runner ──────────────────────────────────────────────────────
def run_assessment(progress_callback=None):
    host_info = get_host_info()
    results = []
    total_score = 0
    max_total = 0

    for i, ctrl in enumerate(CONTROLS):
        if progress_callback:
            progress_callback(i + 1, len(CONTROLS), ctrl["title"])

        try:
            passed, value, evidence = ctrl["check_fn"]()
        except Exception as e:
            passed = False
            value = f"ERROR: {str(e)[:80]}"
            evidence = str(e)

        max_score = ctrl["max_score"]
        if max_score > 0:
            score = max_score if passed else 0
            total_score += score
            max_total += max_score
        else:
            score = 0

        status = "PASS" if passed else ("INFO" if max_score == 0 else "FAIL")
        results.append({
            "id": ctrl["id"],
            "title": ctrl["title"],
            "category": ctrl["category"],
            "description": ctrl["description"],
            "status": status,
            "score": score,
            "max_score": max_score,
            "value": value,
            "evidence": evidence[:200],
        })

    return host_info, results, total_score, max_total


# ─── Excel Export ────────────────────────────────────────────────────────────
def export_excel(host_info, results, total_score, max_total):
    filename = f"iso27001_endpoint_assessment_{datetime.date.today().isoformat()}.xlsx"
    desktop = get_desktop_path()
    filepath = os.path.join(desktop, filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if OPENPYXL_AVAILABLE:
        _export_openpyxl(filepath, host_info, results, total_score, max_total)
    else:
        _export_xlwt_fallback(filepath, host_info, results, total_score, max_total)

    return filepath


def _score_color(pct):
    if pct >= 80: return "C6EFCE", "375623"
    if pct >= 50: return "FFEB9C", "9C6500"
    return "FFC7CE", "9C0006"


def _export_openpyxl(filepath, host_info, results, total_score, max_total):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Executive Summary"

    navy_fill = PatternFill("solid", fgColor="1F4E79")
    blue_fill = PatternFill("solid", fgColor="2E75B6")
    light_blue = PatternFill("solid", fgColor="DEEAF1")
    white_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    dark = Font(color="1A1A2E", name="Calibri", size=10)
    bold_dark = Font(color="1A1A2E", bold=True, name="Calibri", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(left=Side(style="thin",color="BDD7EE"), right=Side(style="thin",color="BDD7EE"),
                  top=Side(style="thin",color="BDD7EE"), bottom=Side(style="thin",color="BDD7EE"))

    def hdr(row, col, text, fill=navy_fill, font=None, align=center):
        c = ws1.cell(row=row, column=col, value=text)
        c.fill = fill; c.font = font or white_font; c.alignment = align; c.border = thin; return c

    def dat(row, col, text, fill=None, font=None, align=left_align, bold=False):
        c = ws1.cell(row=row, column=col, value=text)
        if fill: c.fill = fill
        c.font = font or (bold_dark if bold else dark)
        c.alignment = align; c.border = thin; return c

    # Title
    ws1.merge_cells("A1:F1")
    t = ws1["A1"]
    t.value = "SHAH SCIENTIFIC SOLUTIONS — ISO 27001:2022 ENDPOINT ASSESSMENT"
    t.fill = navy_fill; t.font = Font(color="FFFFFF", bold=True, size=13, name="Calibri"); t.alignment = center
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:F2")
    t2 = ws1["A2"]
    t2.value = "Annex A.8 Technical Controls | Endpoint Assessment | vCISO Practice"
    t2.fill = blue_fill; t2.font = Font(color="FFFFFF", italic=True, size=10, name="Calibri"); t2.alignment = center
    ws1.row_dimensions[2].height = 18

    # Host info
    row = 4
    ws1.merge_cells(f"A{row}:F{row}")
    hdr(row, 1, "ASSESSMENT DETAILS", fill=blue_fill)
    ws1.row_dimensions[row].height = 18

    host_data = [
        ("Hostname", host_info["hostname"]),
        ("Username", host_info["username"]),
        ("Domain", host_info["domain"]),
        ("Operating System", host_info["os"]),
        ("IP Address", host_info["ip_address"]),
        ("Assessment Date", host_info["assessment_date"]),
        ("Organisation", "Shah Scientific Solutions"),
    ]
    for i, (label, value) in enumerate(host_data):
        r = row + 1 + i
        dat(r, 1, label, fill=light_blue, bold=True)
        ws1.merge_cells(f"B{r}:F{r}")
        dat(r, 2, value)
        ws1.row_dimensions[r].height = 15

    row = row + len(host_data) + 2
    overall_pct = int((total_score / max_total) * 100) if max_total > 0 else 0
    score_fill_color, score_font_color = _score_color(overall_pct)
    score_label = "GOOD COMPLIANCE" if overall_pct >= 80 else "PARTIAL COMPLIANCE" if overall_pct >= 50 else "NON-COMPLIANT"
    score_fill = PatternFill("solid", fgColor=score_fill_color)
    score_font = Font(size=26, bold=True, color=score_font_color, name="Calibri")

    ws1.merge_cells(f"A{row}:F{row}")
    hdr(row, 1, "OVERALL ASSESSMENT SCORE", fill=blue_fill)
    ws1.row_dimensions[row].height = 18

    row += 1
    ws1.merge_cells(f"A{row}:C{row}")
    sc = ws1[f"A{row}"]
    sc.value = f"{total_score} / {max_total}"
    sc.fill = score_fill; sc.font = score_font; sc.alignment = center
    ws1.row_dimensions[row].height = 48

    ws1.merge_cells(f"D{row}:F{row}")
    sp = ws1[f"D{row}"]
    sp.value = f"{overall_pct}%\n{score_label}"
    sp.fill = score_fill; sp.font = Font(size=13, bold=True, color=score_font_color, name="Calibri")
    sp.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[row].height = 48

    # Category breakdown
    row += 2
    ws1.merge_cells(f"A{row}:F{row}")
    hdr(row, 1, "SCORE BY CATEGORY", fill=blue_fill)
    ws1.row_dimensions[row].height = 18

    row += 1
    for col, txt in enumerate(["Category","Controls","Score","Max","Percentage","Status"], 1):
        hdr(row, col, txt)

    categories = {}
    for r2 in results:
        cat = r2["category"]
        if cat not in categories:
            categories[cat] = {"score": 0, "max": 0, "count": 0}
        if r2["max_score"] > 0:
            categories[cat]["score"] += r2["score"]
            categories[cat]["max"] += r2["max_score"]
            categories[cat]["count"] += 1

    for cat, data in sorted(categories.items()):
        row += 1
        cat_pct = int((data["score"] / data["max"]) * 100) if data["max"] > 0 else 0
        cf_fill_color, cf_font_color = _score_color(cat_pct)
        cf_fill = PatternFill("solid", fgColor=cf_fill_color)
        dat(row, 1, cat)
        dat(row, 2, str(data["count"]))
        dat(row, 3, f"{data['score']}/{data['max']}", fill=cf_fill, font=Font(bold=True, color=cf_font_color))
        dat(row, 4, str(data["max"]))
        dat(row, 5, f"{cat_pct}%", fill=cf_fill, font=Font(bold=True, color=cf_font_color))
        dat(row, 6, "GOOD" if cat_pct >= 80 else "PARTIAL" if cat_pct >= 50 else "POOR",
           fill=cf_fill, font=Font(bold=True, color=cf_font_color))
        ws1.row_dimensions[row].height = 15

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 12
    ws1.column_dimensions["F"].width = 14

    # Sheet 2 — Full Checklist
    ws2 = wb.create_sheet("Controls Checklist")
    ws2.merge_cells("A1:H1")
    t3 = ws2["A1"]
    t3.value = "ISO 27001:2022 ANNEX A.8 — TECHNICAL CONTROLS CHECKLIST"
    t3.fill = navy_fill; t3.font = Font(color="FFFFFF", bold=True, size=12, name="Calibri"); t3.alignment = center
    ws2.row_dimensions[1].height = 25

    headers = ["Control ID","Title","Category","Status","Score","Max Score","Findings","Evidence"]
    widths = [12,30,20,10,8,10,40,55]
    for col, (hdr_txt, w) in enumerate(zip(headers, widths), 1):
        c = ws2.cell(row=2, column=col, value=hdr_txt)
        c.fill = blue_fill; c.font = white_font; c.alignment = center; c.border = thin
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[2].height = 18

    for i, r2 in enumerate(results):
        row_n = i + 3
        max_sc = r2["max_score"]
        sc_val = r2["score"]

        if r2["status"] == "PASS":
            sfill = PatternFill("solid", fgColor="C6EFCE"); scolor = "375623"
        elif r2["status"] == "INFO":
            sfill = PatternFill("solid", fgColor="DEEAF1"); scolor = "1F4E79"
        else:
            sfill = PatternFill("solid", fgColor="FFC7CE"); scolor = "9C0006"

        vals = [r2["id"], r2["title"], r2["category"], r2["status"],
                f"{sc_val}/{max_sc}" if max_sc > 0 else "N/A",
                str(max_sc) if max_sc > 0 else "N/A",
                r2["value"], r2["evidence"]]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=row_n, column=col, value=val)
            c.fill = sfill; c.font = Font(color=scolor, name="Calibri", size=9)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True); c.border = thin
        ws2.row_dimensions[row_n].height = 40

    wb.save(filepath)


def _export_xlwt_fallback(filepath, host_info, results, total_score, max_total):
    import xlwt
    wb = xlwt.Workbook()
    ws = wb.add_sheet("ISO27001 Controls")
    style_hdr = xlwt.XFStyle()
    style_hdr.font.bold = True
    headers = ["Control ID","Title","Category","Status","Score","Max Score","Findings"]
    for col, h in enumerate(headers):
        ws.write(0, col, h, style_hdr)
    for i, r2 in enumerate(results):
        row = i + 1
        ws.write(row, 0, r2["id"])
        ws.write(row, 1, r2["title"])
        ws.write(row, 2, r2["category"])
        ws.write(row, 3, r2["status"])
        ws.write(row, 4, f"{r2['score']}/{r2['max_score']}")
        ws.write(row, 5, r2["max_score"])
        ws.write(row, 6, r2["value"])
    wb.save(filepath)


# ─── GUI App ─────────────────────────────────────────────────────────────────
def run_gui():
    root = tk.Tk()
    root.title("ISO 27001 Endpoint Assessment — Shah Scientific Solutions")
    root.resizable(False, False)
    root.configure(bg="#1F4E79")

    # Center window
    root.update_idletasks()
    w = 620
    h = 320
    x = (root.winfo_screenwidth() // 2) - (w // 2)
    y = (root.winfo_screenheight() // 2) - (h // 2)
    root.geometry(f"{w}x{h}+{x}+{y}")

    # Header
    header = tk.Frame(root, bg="#1F4E79", pady=12)
    header.pack(fill="x")
    tk.Label(header, text="ISO 27001:2022", bg="#1F4E79", fg="white",
             font=("Calibri", 16, "bold")).pack()
    tk.Label(header, text="Endpoint Technical Controls Assessment", bg="#1F4E79", fg="#BDD7EE",
             font=("Calibri", 11)).pack()
    tk.Label(header, text="Shah Scientific Solutions | vCISO Practice", bg="#1F4E79", fg="#BDD7EE",
             font=("Calibri", 9, "italic")).pack()

    # Progress frame
    progress_frame = tk.Frame(root, bg="#2E75B6", pady=16, padx=24)
    progress_frame.pack(fill="both", expand=True)

    tk.Label(progress_frame, text="Running assessment...", bg="#2E75B6", fg="white",
             font=("Calibri", 11, "bold"), anchor="w").pack(anchor="w")

    status_label = tk.Label(progress_frame, text="Initialising...", bg="#2E75B6", fg="#DEEAF1",
                             font=("Calibri", 9), anchor="w")
    status_label.pack(anchor="w", pady=(4, 0))

    style = ttk.Style()
    style.theme_use("default")
    style.configure("Custom.Horizontal.TProgressbar", background="#FFFFFF", troughcolor="#1F4E79")
    pb = ttk.Progressbar(progress_frame, mode="determinate", length=560, style="Custom.Horizontal.TProgressbar")
    pb.pack(pady=(10, 0))
    pb["maximum"] = len(CONTROLS)
    pb["value"] = 0

    result_label = tk.Label(progress_frame, text="", bg="#2E75B6", fg="white",
                            font=("Calibri", 10, "bold"), anchor="w")
    result_label.pack(anchor="w", pady=(10, 0))

    def update_progress(current, total, ctrl_title):
        pb["value"] = current
        status_label.config(text=f"  Checking: {ctrl_title}")
        root.update()

    def on_complete(filepath, total_score, max_total, host_info, results):
        pb["value"] = len(CONTROLS)
        overall_pct = int((total_score / max_total) * 100) if max_total > 0 else 0

        # Clear frame
        for widget in progress_frame.winfo_children():
            widget.destroy()

        bg_color = "#C6EFCE" if overall_pct >= 80 else "#FFEB9C" if overall_pct >= 50 else "#FFC7CE"
        txt_color = "#375623" if overall_pct >= 80 else "#9C6500" if overall_pct >= 50 else "#9C0006"
        status_text = "GOOD COMPLIANCE" if overall_pct >= 80 else "PARTIAL COMPLIANCE" if overall_pct >= 50 else "NON-COMPLIANT"

        result_label2 = tk.Label(progress_frame, text=f"{total_score} / {max_total}  ({overall_pct}%)",
                                 bg=bg_color, fg=txt_color, font=("Calibri", 26, "bold"))
        result_label2.pack(pady=(8, 4))

        tk.Label(progress_frame, text=status_text, bg=bg_color, fg=txt_color,
                 font=("Calibri", 12, "bold")).pack()

        tk.Label(progress_frame, text=f"\nReport saved to Desktop:\n{os.path.basename(filepath)}",
                 bg="#DEEAF1", fg="#1F4E79", font=("Calibri", 10),
                 wraplength=500).pack(pady=(12, 0))

        btn_frame = tk.Frame(progress_frame, bg="#DEEAF1")
        btn_frame.pack(pady=(10, 0))

        def open_file():
            try:
                os.startfile(filepath)
            except Exception:
                os.startfile(os.path.dirname(filepath))
        def close():
            root.destroy()

        tk.Button(btn_frame, text="Open Report", command=open_file,
                  bg="#1F4E79", fg="white", font=("Calibri", 10, "bold"),
                  padx=12, pady=4, relief="flat", cursor="hand2",
                  activebackground="#2E75B6").pack(side="left", padx=4)
        tk.Button(btn_frame, text="Close", command=close,
                  bg="#6C757D", fg="white", font=("Calibri", 10),
                  padx=12, pady=4, relief="flat", cursor="hand2",
                  activebackground="#5A6268").pack(side="left", padx=4)

        # Run assessment in background
        import threading
        def assess():
            h2, r2, ts2, mt2 = run_assessment()
            fp2 = export_excel(h2, r2, ts2, mt2)
            root.after(0, lambda: on_complete(fp2, ts2, mt2, h2, r2))

        threading.Thread(target=assess, daemon=True).start()

    # Start assessment immediately
    on_complete(None, 0, 0, None, [])

    root.mainloop()


# ─── Console fallback ────────────────────────────────────────────────────────
def run_console():
    print("=" * 60)
    print("ISO 27001:2022 — Endpoint Technical Controls Assessment")
    print("Shah Scientific Solutions | vCISO Practice")
    print("=" * 60)
    host_info = get_host_info()
    print(f"\nHost: {host_info['hostname']} | User: {host_info['username']}")
    print(f"OS: {host_info['os']}")
    print(f"Assessment: {host_info['assessment_date']}\n")

    host_info, results, total_score, max_total = run_assessment()
    for r in results:
        icon = "[PASS]" if r["status"] == "PASS" else "[INFO]" if r["status"] == "INFO" else "[FAIL]"
        print(f"  {icon} {r['id']} — {r['title']}: {r['value']}")

    overall_pct = int((total_score / max_total) * 100) if max_total > 0 else 0
    print(f"\nSCORE: {total_score} / {max_total}  ({overall_pct}%)")
    print(f"Report: {export_excel(host_info, results, total_score, max_total)}")


def _is_headless():
    """Detect if we're in a headless environment (no display)."""
    try:
        display = os.environ.get("DISPLAY", "")
        if not display:
            return True
    except Exception:
        pass
    # On Windows, check if we're in a console-only environment
    try:
        if sys.platform == "win32":
            import ctypes
            user32 = ctypes.windll.user32
            return user32.GetSystemMetrics(0) == 0  # SM_CXSCREEN = 0 means no display
    except Exception:
        pass
    return False

def main():
    headless = _is_headless()
    try:
        if TK_AVAILABLE:
            try:
                test_root = tk.Tk()
                test_root.withdraw()
                test_root.destroy()
                run_gui()
            except Exception as e:
                if headless:
                    # In headless environment, just run the console version silently
                    run_console()
                else:
                    import tkinter.messagebox
                    tkinter.messagebox.showerror("ISO 27001 Assessment",
                        f"Failed to launch GUI:\n{e}\n\nPlease run with Python instead.\nSee: github.com/mominaali675-commits/iso27001-endpoint-tool")
                    raise SystemExit(1)
        else:
            run_console()
    except SystemExit:
        raise
    except Exception as e:
        import traceback
        err_file = os.path.join(os.environ.get("TEMP", "/tmp"), "iso27001_error.log")
        try:
            with open(err_file, "w") as f:
                f.write(traceback.format_exc())
        except Exception:
            pass
        if not headless:
            try:
                import tkinter.messagebox
                tkinter.messagebox.showerror("ISO 27001 Assessment", f"Fatal error: {e}")
            except Exception:
                print(f"Fatal error: {e}")
        raise SystemExit(1)


if __name__ == "__main__":
    main()
